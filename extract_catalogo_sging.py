import openpyxl
import csv
from pathlib import Path
src = Path(r'C:\Users\ximen\OneDrive\Proyectos_DS\SGING\data\raw\Catalogo de Indicadores.xlsx')
dst = Path(r'C:\Users\ximen\OneDrive\Proyectos_DS\Informe_Interactivo\data\objetivos\catalogo_indicadores_sging.csv')
wb = openpyxl.load_workbook(src, data_only=True)
ws = wb['Catalogo Indicadores']
headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
cols = ['Id','Indicador','Clasificacion','Linea_Estrategica','Objetivo_Estrategico','Meta_Estrategica','Plan_Anual','Indicadores_Clave','Indicadores_Plan_Estrategico','Indicadores_Vicerrectoria','Subprocesos','General','Ind_Act']
indexes = [headers.index(c) for c in cols]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rows.append([row[i] for i in indexes])
with dst.open('w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(cols)
    writer.writerows(rows)
print('wrote', dst)
