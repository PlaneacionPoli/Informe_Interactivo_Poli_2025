import openpyxl
from pathlib import Path
path = Path(r"C:\Users\ximen\OneDrive\Proyectos_DS\SGING\data\raw\Catalogo de Indicadores.xlsx")
print('exists', path.exists(), 'size', path.stat().st_size if path.exists() else 'n/a')
wb = openpyxl.load_workbook(path, data_only=False)
print('sheets', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    print('sheet', name)
    for r in rows:
        print(r)
