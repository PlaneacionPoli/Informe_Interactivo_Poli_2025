import openpyxl
from pathlib import Path
path = Path(r'C:\Users\ximen\OneDrive\Proyectos_DS\SGING\data\raw\Catalogo de Indicadores.xlsx')
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Catalogo Indicadores']
headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
print('HEADERS:', headers)
cols = ['Id','Indicador','Clasificacion','Linea_Estrategica','Objetivo_Estrategico','Meta_Estrategica','Plan_Anual','Indicadores_Clave','Indicadores_Plan_Estrategico','Indicadores_Vicerrectoria','Subprocesos','General','Ind_Act']
indexes = {c: headers.index(c) for c in cols if c in headers}
print('INDEXES:', indexes)
print('\nSAMPLE ROWS:')
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=12, values_only=True), start=2):
    data = {col: row[idx] for col, idx in indexes.items()}
    print(i, data)
